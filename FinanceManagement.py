from unicodedata import decimal
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

#variable allows for one to change the location of spreadsheet to a different one within this file
default = '/Users/alaincourtines/Desktop/Projects/Project II/W2.xlsx'
#creates an object of an already made excel worksheet
book = load_workbook(default)
#following variable gives one access to the worksheet present in this workbook
sheet = book.active




def summer(x,y):
    return x + y

print("\nWELCOME TO YOUR EXPENSE TRACKER, THE ONE STOP SHOP FOR TRACKING ALL EXPENSES \n")
print("the options for type of expense entry come as follow:")
print("{'food', 'drink', 'gas', 'transportation', 'social', 'groceries', and 'living'} \n")
print("the options for inquiry into your spending are as follow:")
print("{'total spent', 'food spending', 'drink spending', 'transport spending',")
print("'social spending', 'grocery spending', and 'living expenses'}\n")

print("\nthere is also an option to clear all current numerical information")
print("from your excel file if you like. Just enter the command 'clear'")

print("if you ever want to quit talking to me just type 'quit' in the terminal\n")




IP = input("alrighty, what'd you spend on or what do you want to know about now? \n")




while(IP is not "quit"):

    if IP == "quit":
        print("see ya later boss")
        break
    elif IP == "clear":
        for column in range(1,8):
            for row in range(2,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if not (sheet[cell].value is None):
                    sheet[cell] = None
                                
        book.save(default)
        print("you got a blank slate now boss")
        break
    elif IP == "food" or IP == "drink":
        amount = input("how much did you spend?\n")
        for column in range(1,2):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if(sheet[cell].value is None):
                    sheet[cell] = float(amount)
                    break
                    
        book.save(default)
        print("dully noted")
        break
        
    elif IP == "gas" or IP == "transportation":
        
        amount = input("how much did you spend?\n")
        for column in range(2,3):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if(sheet[cell].value is None):
                    sheet[cell] = float(amount)
                    break
                    
        book.save(default)
        print("got it boss")
        break

    elif IP == "social":
        amount = input("how much did you spend?\n")
        for column in range(3,4):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if(sheet[cell].value is None):
                    sheet[cell] = float(amount)
                    break
                    
        book.save(default)
        print("hope it was worth it boss")
        break
    elif IP == "groceries":
        amount = input("how much did you spend?\n")
        for column in range(4,5):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if(sheet[cell].value is None):
                    sheet[cell] = float(amount)
                    break
        book.save(default)
        print("go cook something nice for yourself boss")
        break
    elif IP == "living":
        amount = input("how much did you spend?\n")
        for column in range(5,6):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if(sheet[cell].value is None):
                    sheet[cell] = float(amount)
                    break
        book.save(default)
        print("for sure")
        break
    elif IP == "food spending" or IP == "drink spending":
        amount = 0
        for column in range(1,2):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if not (sheet[cell].value is None):
                    if(type(sheet[cell].value) == int or type(sheet[cell].value) == float):
                        amount = summer(amount, float(sheet[cell].value))
        sheet['F2'] = amount
        book.save(default)
        print("\nyou've spent $"+ str(round((sheet['F2'].value),4)) + " on food and drink as of recent\n")
        break
    elif IP == "transport spending":
        amount = 0
        for column in range(2,3):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if not (sheet[cell].value is None):
                    if(type(sheet[cell].value) == int or type(sheet[cell].value) == float):
                        amount = summer(amount, float(sheet[cell].value))
        sheet['F3'] = amount
        book.save(default)
        print("\nyou've spent $"+ str(round((sheet['F3'].value),4)) + " on transportation as of recent\n")
        break
    elif IP == "social spending":
        amount = 0
        for column in range(3,4):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if not (sheet[cell].value is None):
                    if(type(sheet[cell].value) == int or type(sheet[cell].value) == float):
                        amount = summer(amount, float(sheet[cell].value))
        sheet['F4'] = amount
        book.save(default)
        print("\nyou've spent $"+ str(round((sheet['F4'].value),4)) + " on your social life as of recent\n")
        break
    elif IP == "grocery spending":
        amount = 0
        for column in range(4,5):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if not (sheet[cell].value is None):
                    if(type(sheet[cell].value) == int or type(sheet[cell].value) == float):
                        amount = summer(amount, float(sheet[cell].value))
        sheet['F5'] = amount
        book.save(default)
        print("\nyou've spent $"+ str(round((sheet['F5'].value),4)) + " on groceries as of recent\n")
        break
    elif IP == "living expenses":
        amount = 0
        for column in range(5,6):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if not (sheet[cell].value is None):
                    if(type(sheet[cell].value) == int or type(sheet[cell].value) == float):
                        amount = summer(amount, float(sheet[cell].value))
        sheet['F6'] = amount
        book.save(default)
        print("\nyou've spent $"+ str(round((sheet['F6'].value),4)) + " on living expenses as of recent\n")
        break
    elif IP == "total spent":
        amount = 0
        for column in range(1,6):
            for row in range(1,1000):
                letter = get_column_letter(column)
                cell = letter + str(row)
                if not (sheet[cell].value is None):
                    if(type(sheet[cell].value) == int or type(sheet[cell].value) == float):
                        amount = summer(amount, float(sheet[cell].value))
        sheet['G2'] = amount
        book.save(default)
        print("\nyou've spent $"+ str(round((sheet['G2'].value),4)) + " on living expenses as of recent\n")
        break
    else:
        print("that's not an option boss, try again")
        break

